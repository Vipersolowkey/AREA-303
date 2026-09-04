"""Parsing, validation and atomic commit for seller-provided tabular data."""

from __future__ import annotations

import csv
import io
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import BusinessRuleError, NotFoundError, ValidationError
from app.models.marketplace import ShopOrder, ShopProduct
from app.models.onboarding_data import (
    WorkspaceDataImport,
    WorkspaceDataImportRow,
    WorkspaceDataRecord,
    WorkspaceMarketplaceAccount,
)
from app.services import marketplace_link

MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 10_000
DATASET_FIELDS = {
    "products": ("sku", "name", "price", "stock", "category"),
    "orders": ("order_id", "ordered_at", "total_amount", "status"),
}
REQUIRED_FIELDS = {
    "products": ("sku", "name", "price", "stock"),
    "orders": ("order_id", "ordered_at", "total_amount"),
}


def _cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def parse_tabular_file(filename: str, content: bytes) -> tuple[str, list[str], list[dict[str, str]]]:
    """Return a bounded rectangular table; no row is persisted at this stage."""
    if not content:
        raise ValidationError("Tệp trống.")
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValidationError("Tệp vượt quá giới hạn 10 MB.")
    suffix = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if suffix == "csv":
        text = next((content.decode(encoding) for encoding in ("utf-8-sig", "utf-8", "cp1258") if _can_decode(content, encoding)), None)
        if text is None:
            raise ValidationError("CSV phải dùng mã hóa UTF-8 hoặc Windows-1258.")
        raw = list(csv.reader(io.StringIO(text)))
        kind = "csv"
    elif suffix in {"xlsx", "xlsm"}:
        try:
            book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = book.active
            raw = [[_cell(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
            book.close()
        except Exception as exc:  # openpyxl exposes several format-specific errors
            raise ValidationError("Không đọc được file Excel. Hãy dùng .xlsx hợp lệ.") from exc
        kind = "xlsx"
    else:
        raise ValidationError("Chỉ hỗ trợ tệp .csv, .xlsx hoặc .xlsm.")
    if not raw:
        raise ValidationError("Tệp không có dòng tiêu đề.")
    headers = [_cell(value) for value in raw[0]]
    if not headers or any(not header for header in headers):
        raise ValidationError("Dòng tiêu đề không được để trống.")
    if len(set(headers)) != len(headers):
        raise ValidationError("Tên cột bị trùng. Hãy đổi tên từng cột trước khi import.")
    rows = [
        {header: _cell(values[index]) if index < len(values) else "" for index, header in enumerate(headers)}
        for values in raw[1:]
        if any(_cell(value) for value in values)
    ]
    if not rows:
        raise ValidationError("Tệp chỉ có tiêu đề, chưa có dữ liệu.")
    if len(rows) > MAX_IMPORT_ROWS:
        raise ValidationError(f"Mỗi lần import tối đa {MAX_IMPORT_ROWS:,} dòng.")
    return kind, headers, rows


def _can_decode(content: bytes, encoding: str) -> bool:
    try:
        content.decode(encoding)
        return True
    except UnicodeDecodeError:
        return False


def _number(value: str, field: str, *, integer: bool = False) -> int | None:
    normalized = value.strip().replace(" ", "").replace("₫", "").replace("VND", "")
    normalized = normalized.replace(".", "").replace(",", ".") if "," in normalized and "." not in normalized else normalized.replace(",", "")
    try:
        number = Decimal(normalized)
    except InvalidOperation:
        return None
    if number < 0 or (integer and number != number.to_integral_value()):
        return None
    return int(number)


def validate_row(dataset_type: str, raw: dict[str, str], mapping: dict[str, str]) -> tuple[dict[str, object], list[str]]:
    normalized: dict[str, object] = {
        field: raw.get(column, "").strip() for field, column in mapping.items() if column
    }
    errors: list[str] = []
    for field in REQUIRED_FIELDS[dataset_type]:
        if not normalized.get(field):
            errors.append(f"Thiếu {field}.")
    if dataset_type == "products":
        if normalized.get("sku") and len(str(normalized["sku"])) > 160:
            errors.append("SKU dài quá 160 ký tự.")
        if normalized.get("name") and len(str(normalized["name"])) > 512:
            errors.append("Tên sản phẩm dài quá 512 ký tự.")
        for field, integer in (("price", False), ("stock", True)):
            if normalized.get(field):
                number = _number(str(normalized[field]), field, integer=integer)
                if number is None:
                    errors.append(f"{field} phải là số không âm" + (" nguyên." if integer else "."))
                else:
                    normalized[field] = number
    else:
        if normalized.get("order_id") and len(str(normalized["order_id"])) > 160:
            errors.append("Mã đơn dài quá 160 ký tự.")
        if normalized.get("total_amount"):
            number = _number(str(normalized["total_amount"]), "total_amount")
            if number is None:
                errors.append("total_amount phải là số không âm.")
            else:
                normalized["total_amount"] = number
        if normalized.get("ordered_at"):
            try:
                parsed = datetime.fromisoformat(str(normalized["ordered_at"]).replace("Z", "+00:00"))
                normalized["ordered_at"] = (parsed if parsed.tzinfo else parsed.replace(tzinfo=UTC)).isoformat()
            except ValueError:
                errors.append("ordered_at phải là ngày giờ ISO, ví dụ 2026-09-05T10:30:00+07:00.")
    return normalized, errors


async def create_preview(db: AsyncSession, *, workspace_id: int, user_id: int, dataset_type: str, filename: str, content: bytes) -> WorkspaceDataImport:
    if dataset_type not in DATASET_FIELDS:
        raise ValidationError("Loại dữ liệu import không hợp lệ.")
    kind, headers, rows = parse_tabular_file(filename, content)
    batch = WorkspaceDataImport(
        workspace_id=workspace_id, created_by_user_id=user_id, dataset_type=dataset_type,
        filename=filename[:255], file_kind=kind, headers=headers, total_rows=len(rows),
    )
    db.add(batch)
    await db.flush()
    db.add_all(WorkspaceDataImportRow(import_id=batch.id, row_number=index + 2, raw_values=row) for index, row in enumerate(rows))
    await db.commit()
    await db.refresh(batch)
    return batch


async def get_import(db: AsyncSession, *, workspace_id: int, import_id: int) -> WorkspaceDataImport:
    row = await db.get(WorkspaceDataImport, import_id)
    if row is None or row.workspace_id != workspace_id:
        raise NotFoundError("Không tìm thấy phiên import trong workspace.")
    return row


async def list_rows(db: AsyncSession, import_id: int) -> list[WorkspaceDataImportRow]:
    result = await db.execute(select(WorkspaceDataImportRow).where(WorkspaceDataImportRow.import_id == import_id).order_by(WorkspaceDataImportRow.row_number))
    return list(result.scalars().all())


async def validate_import(db: AsyncSession, *, batch: WorkspaceDataImport, mapping: dict[str, str]) -> list[WorkspaceDataImportRow]:
    allowed = set(DATASET_FIELDS[batch.dataset_type])
    if any(field not in allowed for field in mapping) or any(column not in batch.headers for column in mapping.values()):
        raise ValidationError("Mapping có cột hoặc trường không hợp lệ.")
    missing = [field for field in REQUIRED_FIELDS[batch.dataset_type] if not mapping.get(field)]
    if missing:
        raise ValidationError(f"Cần map các trường bắt buộc: {', '.join(missing)}.")
    if len(set(mapping.values())) != len(mapping.values()):
        raise ValidationError("Một cột chỉ được map vào một trường.")
    rows = await list_rows(db, batch.id)
    for row in rows:
        normalized, errors = validate_row(batch.dataset_type, row.raw_values, mapping)
        row.normalized_values = normalized
        row.errors = errors
        row.is_valid = not errors
    batch.mapping = mapping
    batch.valid_rows = sum(row.is_valid for row in rows)
    batch.invalid_rows = len(rows) - batch.valid_rows
    batch.status = "validated"
    await db.commit()
    return rows


async def confirm_import(db: AsyncSession, *, batch: WorkspaceDataImport) -> int:
    if batch.status != "validated" or not batch.mapping:
        raise BusinessRuleError("Hãy map cột và chạy validation trước khi xác nhận.")
    if batch.invalid_rows:
        raise BusinessRuleError("Hãy sửa mọi lỗi theo dòng trước khi ghi dữ liệu vào workspace.")
    if not batch.valid_rows:
        raise BusinessRuleError("Không có dòng hợp lệ để ghi.")
    rows = await list_rows(db, batch.id)
    for row in rows:
        payload = row.normalized_values or {}
        external_key = str(payload["sku"] if batch.dataset_type == "products" else payload["order_id"])
        found = await db.execute(select(WorkspaceDataRecord).where(
            WorkspaceDataRecord.workspace_id == batch.workspace_id,
            WorkspaceDataRecord.dataset_type == batch.dataset_type,
            WorkspaceDataRecord.external_key == external_key,
        ))
        record = found.scalar_one_or_none()
        if record is None:
            db.add(WorkspaceDataRecord(workspace_id=batch.workspace_id, import_id=batch.id, dataset_type=batch.dataset_type, external_key=external_key, payload=payload))
        else:
            record.import_id = batch.id
            record.payload = payload
    batch.status = "committed"
    await db.commit()
    return batch.valid_rows


async def readiness(db: AsyncSession, workspace_id: int) -> dict[str, object]:
    imported = await db.scalar(select(func.count()).select_from(WorkspaceDataRecord).where(WorkspaceDataRecord.workspace_id == workspace_id)) or 0
    account_id = await db.scalar(select(WorkspaceMarketplaceAccount.seller_account_id).where(WorkspaceMarketplaceAccount.workspace_id == workspace_id))
    marketplace_records = 0
    shops: list[dict[str, object]] = []
    if account_id:
        # ShopProduct and ShopOrder are attached to ShopConnection, which is
        # deliberately the only bridge to the OAuth seller account.
        from app.models.marketplace import ShopConnection
        product_count = await db.scalar(select(func.count()).select_from(ShopProduct).join(ShopConnection).where(ShopConnection.seller_account_id == account_id)) or 0
        order_count = await db.scalar(select(func.count()).select_from(ShopOrder).join(ShopConnection).where(ShopConnection.seller_account_id == account_id)) or 0
        marketplace_records = product_count + order_count
        result = await db.execute(select(ShopConnection).where(ShopConnection.seller_account_id == account_id))
        shops = [{"platform": row.platform, "shop_name": row.shop_name, "status": row.status, "last_synced_at": row.last_synced_at} for row in result.scalars().all()]
    total = int(imported) + int(marketplace_records)
    return {"ready": total > 0, "total_records": total, "manual_records": int(imported), "marketplace_records": int(marketplace_records), "shops": shops}


async def list_workspace_products(db: AsyncSession, workspace_id: int) -> list[dict[str, object]]:
    """The confirmed import catalogue, never the public/demo storefront."""
    result = await db.execute(
        select(WorkspaceDataRecord)
        .where(
            WorkspaceDataRecord.workspace_id == workspace_id,
            WorkspaceDataRecord.dataset_type == "products",
        )
        .order_by(WorkspaceDataRecord.updated_at.desc())
    )
    return [
        {
            "id": row.external_key,
            "sku": str(row.payload.get("sku", row.external_key)),
            "name": str(row.payload.get("name", "")),
            "price_vnd": int(row.payload.get("price", 0)),
            "stock": int(row.payload.get("stock", 0)),
            "category": str(row.payload.get("category", "Chưa phân loại")),
        }
        for row in result.scalars().all()
    ]


async def begin_marketplace_connection(db: AsyncSession, *, workspace_id: int, user_id: int, workspace_name: str, platform: str) -> str:
    mapping = await db.scalar(select(WorkspaceMarketplaceAccount).where(WorkspaceMarketplaceAccount.workspace_id == workspace_id))
    if mapping is None:
        account = await marketplace_link.create_seller_account(db, name=workspace_name, user_id=user_id)
        mapping = WorkspaceMarketplaceAccount(workspace_id=workspace_id, seller_account_id=account.id)
        db.add(mapping)
        await db.commit()
    return await marketplace_link.begin_authorisation(db, mapping.seller_account_id, platform)


async def workspace_for_seller_account(db: AsyncSession, seller_account_id: int) -> int | None:
    return await db.scalar(select(WorkspaceMarketplaceAccount.workspace_id).where(WorkspaceMarketplaceAccount.seller_account_id == seller_account_id))
